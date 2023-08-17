import openpyxl

# File--> Workbook-->Sheet--> Rows--> Cells

file = "D:\SeleniumWithPythonPractice\data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row  # count number of rows in an Excel sheet 6
cols = sheet.max_column  # count number of columns in an Excel sheet 4

# Reading all the rows and columns from an Excel sheet
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value, end='    ')
    print()

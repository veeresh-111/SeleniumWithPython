import openpyxl

# same data
# file = "D:\SeleniumWithPythonPractice\writedata.xlsx"
#
# workbook = openpyxl.load_workbook(file)
# sheet = workbook.active  # get active sheet means if single sheet in excel
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r, c).value="welcome"
#
# workbook.save(file)  # save the file after enter the data into excel

# Multiple data
file = 'D:\\SeleniumWithPythonPractice\\testdata.xlsx'

workbook = openpyxl.load_workbook(file)
sheet = workbook.active  # get active sheet means if single sheet in excel

sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="engineer"

sheet.cell(2,1).value=567
sheet.cell(2,2).value="john"
sheet.cell(2,3).value="doctor"

sheet.cell(3,1).value=243
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="salesman"

workbook.save(file)  # save the file after enter the data into excel